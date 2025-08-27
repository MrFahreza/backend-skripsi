from flask import jsonify, current_app, send_file
from io import BytesIO
import pandas as pd
from . import saw_bp
from ..utils.decorators import token_required
from ..utils.email_utils import send_saw_warning_email, send_saw_congrats_email
import numpy as np
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# --- Konfigurasi Bobot Kriteria ---
BOBOT_SAW = {
    "c1": 0.25, # Keaktifan Organisasi
    "c2": 0.35, # IPK
    "c3": 0.40  # Persentase Kehadiran
}

# --- Helper Function untuk Rating Kecocokan ---
def _get_rating(c1, c2, c3):
    """Mengubah nilai kriteria asli menjadi skala rating 1-5."""
    # Rating C1: Keaktifan Organisasi
    if c1 == 5: r1 = 5
    elif c1 >= 3: r1 = 4
    elif c1 >= 2: r1 = 3
    elif c1 >= 1: r1 = 2
    else: r1 = 1
    
    # Rating C2: IPK
    if c2 == 4: r2 = 5
    elif c2 >= 2.75: r2 = 4
    elif c2 >= 2: r2 = 3
    elif c2 >= 1: r2 = 2
    else: r2 = 1
    
    # Rating C3: Persentase Kehadiran (diasumsikan 0-1)
    if c3 >= 1.0: r3 = 5
    elif c3 >= 0.85: r3 = 4
    elif c3 >= 0.75: r3 = 3
    elif c3 >= 0.60: r3 = 2
    else: r3 = 1
    
    return r1, r2, r3

# --- Proses Inti Perhitungan Simple Additive Weighting ---
def _run_saw_calculation_logic(app, period_name="Manual"):
    with app.app_context():
        penilaian_collection = current_app.db.penilaian_mahasiswa
        mahasiswa_collection = current_app.db.mahasiswa
        
        all_penilaian = list(penilaian_collection.find({}))
        if not all_penilaian:
            return {"success": False, "message": "Tidak ada data penilaian untuk dihitung"}
            
        # --- Mengambil Data Penilaian Mahasiswa ---
        student_emails = {m['npm']: m['email'] for m in mahasiswa_collection.find({}, {'npm': 1, 'email': 1})}

        # --- Membuat Matriks Keputusan Data ---
        matriks_x = []
        for p in all_penilaian:
            c1 = p.get('keaktifan_organisasi', 0)
            c2 = p.get('ipk', 0)
            c3 = p.get('persentase_kehadiran', 0)
            r1, r2, r3 = _get_rating(c1, c2, c3)
            matriks_x.append({
                "npm": p['npm'], "nama": p['nama'],
                "c1_rated": r1, "c2_rated": r2, "c3_rated": r3
            })
        
        # --- Menentukan Nilai Max ---
        max_c1_relatif = max((item['c1_rated'] for item in matriks_x), default=1)
        max_c2_relatif = max((item['c2_rated'] for item in matriks_x), default=1)
        max_c3_relatif = max((item['c3_rated'] for item in matriks_x), default=1)
        MAX_STANDAR = 5.0

        hasil_akhir = []
        mail_config = {
            "MAIL_SERVER": current_app.config['MAIL_SERVER'],
            "MAIL_PORT": current_app.config['MAIL_PORT'],
            "MAIL_USERNAME": current_app.config['MAIL_USERNAME'],
            "MAIL_PASSWORD": current_app.config['MAIL_PASSWORD']
        }

        
        for i, item_x in enumerate(matriks_x):
            original_assessment = all_penilaian[i]

            # --- Normalisasi dan menghitung skor akhir SAW ---
            r1_relatif = item_x['c1_rated'] / max_c1_relatif
            r2_relatif = item_x['c2_rated'] / max_c2_relatif
            r3_relatif = item_x['c3_rated'] / max_c3_relatif

            skor_akhir_saw = (
                (r1_relatif * BOBOT_SAW['c1']) +
                (r2_relatif * BOBOT_SAW['c2']) +
                (r3_relatif * BOBOT_SAW['c3'])
            )

            # --- Normalisasi dan menghitung skor akhir Standar ---
            r1_standar = item_x['c1_rated'] / MAX_STANDAR
            r2_standar = item_x['c2_rated'] / MAX_STANDAR
            r3_standar = item_x['c3_rated'] / MAX_STANDAR
            skor_akhir_standar = (
                (r1_standar * BOBOT_SAW['c1']) +
                (r2_standar * BOBOT_SAW['c2']) +
                (r3_standar * BOBOT_SAW['c3'])
            )
            
            student_email = student_emails.get(item_x['npm'])
            status = "Standar Terpenuhi"
            
            # --- Menentukan status peringatan <0.7 ---
            if student_email:
                if skor_akhir_standar < 0.7:
                    status = "Perlu Peringatan"
                    kriteria_lemah = []
                    if item_x['c1_rated'] <= 3: kriteria_lemah.append("Keaktifan Organisasi")
                    if item_x['c2_rated'] < 4: kriteria_lemah.append("IPK")
                    if item_x['c3_rated'] <= 2: kriteria_lemah.append("Persentase Kehadiran")
                    
                    # --- Mengirimkan email peringatan ---
                    if kriteria_lemah:
                        send_saw_warning_email(student_email, item_x['nama'], kriteria_lemah, original_assessment, mail_config, "ini")
                else:
                    # --- Mengirimkan email ucapan selamat ---
                    send_saw_congrats_email(student_email, item_x['nama'], original_assessment, mail_config, "ini")

            # --- Menyimpan hasil data array data sementara
            hasil_akhir.append({
                "npm": item_x['npm'],
                "nama": item_x['nama'],
                "c1_normalized": round(r1_relatif, 3),
                "c2_normalized": round(r2_relatif, 3),
                "c3_normalized": round(r3_relatif, 3),
                "skor_akhir_saw": round(skor_akhir_saw, 4),
                "skor_akhir_standar": round(skor_akhir_standar, 4),
                "status": status
            })
        
        # --- Mengurutkan hasil data perhitungan ---
        hasil_saw_sorted = sorted(hasil_akhir, key=lambda x: x['skor_akhir_saw'], reverse=True)
        for i, item in enumerate(hasil_saw_sorted):
            item['ranking_saw'] = i + 1

        hasil_standar_sorted = sorted(hasil_akhir, key=lambda x: x['skor_akhir_standar'], reverse=True)
        standar_rank_map = {item['npm']: i + 1 for i, item in enumerate(hasil_standar_sorted)}

        for item in hasil_saw_sorted:
            item['ranking_standar'] = standar_rank_map.get(item['npm'])

        # --- Menghapus data hasil perhitungan yang lama dari database ---
        hasil_collection = current_app.db.hasil_penilaian_saw
        hasil_collection.delete_many({})
        if hasil_saw_sorted:
            # --- Menyimpan data hasil perhitungan yang baru ke database
            hasil_collection.insert_many(hasil_saw_sorted)
            
        # --- Menampilkan notifikasi perthiungan SAW berhasil dilakukan
        print("LOG: Perhitungan SAW selesai dijalankan oleh helper.")
        return {"success": True, "message": "Perhitungan SAW berhasil"}
    
# --- Endpoint Untuk Melakukan Perhitungan Menggunakan Metode SAW ---
@saw_bp.route('/calculate', methods=['POST'])
@token_required
def calculate_saw(current_user_id):
    result = _run_saw_calculation_logic(current_app._get_current_object())
    
    if result["success"]:
        return jsonify({"code": 200, "message": "Perhitungan SAW berhasil dan hasil telah disimpan."}), 200
    else:
        return jsonify({"code": 404, "message": result["message"]}), 404

# --- Endpoint untuk Melihat Hasil Perhitungan SAW ---
@saw_bp.route('/', methods=['GET'])
@token_required
def get_saw_results(current_user_id):
    hasil_collection = current_app.db.hasil_penilaian_saw
    results = list(hasil_collection.find({}, {'_id': 0}))
    if not results:
        return jsonify({"code": 404, "message": "Hasil perhitungan belum ada. Silakan picu perhitungan terlebih dahulu."}), 404
    return jsonify({"code": 200, "data": results}), 200

# --- Endpoint untuk Export Hasil Perhitungan SAW ---
@saw_bp.route('/export', methods=['GET'])
@token_required
def export_saw_results(current_user_id):
    hasil_collection = current_app.db.hasil_penilaian_saw
    results = list(hasil_collection.find({}, {'_id': 0}))

    if not results:
        return jsonify({"code": 404, "message": "Tidak ada data hasil perhitungan untuk diekspor."}), 404
        
    df = pd.DataFrame(results)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Hasil', startrow=2, header=False)

        workbook = writer.book
        worksheet = writer.sheets['Hasil']

        # Styles Text
        font_title = Font(name='Times New Roman', size=12, bold=True)
        font_header = Font(name='Times New Roman', size=12, bold=True)
        font_data = Font(name='Times New Roman', size=12, bold=False)
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # =Style Judul Utama
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = worksheet['A1']
        title_cell.value = "HASIL AKHIR PENILIAN KINERJA MAHASISWA DENGAN METODE SAW"
        title_cell.font = font_title
        title_cell.alignment = center_align

        # Style Header Kolom
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=2, column=col_num)
            cell.value = column_title.replace('_', ' ').title()
            cell.font = font_header
            cell.border = thin_border
            cell.alignment = center_align

        # Style Data dan Border
        for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.font = font_data
                cell.border = thin_border
        
        # Lebar Kolom
        for col_num, column_title in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(column_title)
            for cell in worksheet[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    
        
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 35

    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Hasil Penilaian SAW.xlsx'
    )

# --- Endpoint untuk mengirim peringatan manual ---
@saw_bp.route('/send-warning/<npm>', methods=['POST'])
@token_required
def send_single_warning(current_user_id, npm):
    mahasiswa_collection = current_app.db.mahasiswa
    penilaian_collection = current_app.db.penilaian_mahasiswa

    # 1. Ambil data mahasiswa dan data penilaian berdasarkan NPM
    mahasiswa_data = mahasiswa_collection.find_one({"npm": npm})
    penilaian_data = penilaian_collection.find_one({"npm": npm})

    if not mahasiswa_data:
        return jsonify({"code": 404, "message": f"Mahasiswa dengan NPM {npm} tidak ditemukan."}), 404
    if not penilaian_data:
        return jsonify({"code": 404, "message": f"Data penilaian untuk NPM {npm} tidak ditemukan."}), 404

    # 2. Identifikasi kriteria yang lemah
    c1 = penilaian_data.get('keaktifan_organisasi', 0)
    c2 = penilaian_data.get('ipk', 0)
    c3 = penilaian_data.get('persentase_kehadiran', 0)
    r1, r2, r3 = _get_rating(c1, c2, c3)

    kriteria_lemah = []
    if r1 <= 2: kriteria_lemah.append("Keaktifan Organisasi")
    if r2 <= 3: kriteria_lemah.append("IPK")
    if r3 <= 2: kriteria_lemah.append("Persentase Kehadiran")

    # 3. Cek apakah ada kriteria yang lemah untuk dilaporkan
    if not kriteria_lemah:
        return jsonify({"code": 200, "message": "Mahasiswa ini tidak memiliki kriteria di bawah standar untuk diperingatkan."}), 200
        
    # 4. Kirim email
    mail_config = {
        "MAIL_SERVER": current_app.config['MAIL_SERVER'],
        "MAIL_PORT": current_app.config['MAIL_PORT'],
        "MAIL_USERNAME": current_app.config['MAIL_USERNAME'],
        "MAIL_PASSWORD": current_app.config['MAIL_PASSWORD']
    }
    
    email_sent = send_saw_warning_email(
        mahasiswa_data['email'], 
        mahasiswa_data['nama'], 
        kriteria_lemah, 
        penilaian_data,
        mail_config,
        ""
    )

    if email_sent:
        return jsonify({"code": 200, "message": f"Email peringatan berhasil dikirim ke {mahasiswa_data['nama']}."}), 200
    else:
        return jsonify({"code": 500, "message": "Gagal mengirim email peringatan."}), 500